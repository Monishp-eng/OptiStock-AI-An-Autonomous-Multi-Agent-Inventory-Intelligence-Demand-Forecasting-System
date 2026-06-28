"""
export_router.py — PDF and Excel inventory export endpoints
"""
import io
from datetime import date
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse

from src.database import get_db
from src.auth import get_current_user

router = APIRouter(prefix="/api/export", tags=["Export"])


@router.get("/excel")
def export_excel(db=Depends(get_db),
                 current_user: dict = Depends(get_current_user)):
    """Export full inventory as Excel (.xlsx) with two sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # ── Sheet 1: Inventory ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Inventory"

    headers = ["SKU", "Product Name", "Category", "Current Stock",
               "Cost/Unit (₹)", "Selling Price (₹)", "Supplier",
               "Lead Time (days)", "Defect Rate"]
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    items = db.execute("""
        SELECT sku, product_name, category, current_stock, cost_per_unit,
               selling_price, supplier, lead_time_days, defect_rate
        FROM inventory_items ORDER BY sku
    """).fetchall()

    for row_idx, item in enumerate(items, 2):
        for col_idx, val in enumerate(item, 1):
            ws1.cell(row=row_idx, column=col_idx, value=val)

    for col in ws1.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    # ── Sheet 2: Sales (last 30 days) ───────────────────────────────────────
    ws2 = wb.create_sheet("Sales (30 days)")
    headers2 = ["Date", "SKU", "Product Name", "Qty Sold"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    sales = db.execute("""
        SELECT s.sale_date, s.sku, i.product_name, s.quantity
        FROM daily_sales s
        LEFT JOIN inventory_items i ON s.sku = i.sku
        WHERE s.sale_date >= date('now', '-30 days')
        ORDER BY s.sale_date DESC
    """).fetchall()

    for row_idx, sale in enumerate(sales, 2):
        for col_idx, val in enumerate(sale, 1):
            ws2.cell(row=row_idx, column=col_idx, value=val)

    # Stream the workbook
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"optistock_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/pdf")
def export_pdf(db=Depends(get_db),
               current_user: dict = Depends(get_current_user)):
    """Export inventory as a styled PDF table."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph(f"OptiStock — Inventory Report ({date.today()})",
                               styles["Title"]))
    elements.append(Spacer(1, 0.5*cm))

    # Table data
    col_headers = ["SKU", "Product Name", "Category", "Stock",
                   "Cost/Unit", "Selling Price", "Supplier", "Lead Time"]
    items = db.execute("""
        SELECT sku, product_name, category, current_stock, cost_per_unit,
               selling_price, supplier, lead_time_days
        FROM inventory_items ORDER BY sku
    """).fetchall()

    data = [col_headers]
    for item in items:
        row = list(item)
        row[4] = f"₹{row[4]:,.2f}"
        row[5] = f"₹{row[5]:,.2f}" if row[5] else "—"
        row[7] = f"{row[7]}d"
        data.append(row)

    col_widths = [2.5*cm, 5.5*cm, 3*cm, 2*cm, 2.8*cm, 3*cm, 4.5*cm, 2.2*cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0),  9),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME",    (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",    (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4FF")]),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("TOPPADDING",  (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    doc.build(elements)
    buf.seek(0)

    filename = f"optistock_inventory_{date.today().isoformat()}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
